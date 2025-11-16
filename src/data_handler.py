"""
Data Handler Module
Manages reading and writing application portfolio data from/to CSV and Excel files.
"""

import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Union
import logging
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.chart import BarChart, PieChart, ScatterChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
import numpy as np

logger = logging.getLogger(__name__)


class DataHandler:
    """
    Handles all data I/O operations for the application rationalization tool.

    Supports CSV and Excel formats with validation and error handling.
    """

    REQUIRED_COLUMNS = [
        'Application Name',
        'Owner',
        'Business Value',
        'Tech Health',
        'Cost',
        'Usage',
        'Security',
        'Strategic Fit',
        'Redundancy'
    ]

    OUTPUT_COLUMNS = REQUIRED_COLUMNS + [
        'Composite Score',
        'Action Recommendation',
        'Comments',
        'TIME Category',
        'TIME Rationale',
        'TIME Business Value Score',
        'TIME Technical Quality Score'
    ]

    def __init__(self, data_dir: Optional[Path] = None):
        """
        Initialize the data handler.

        Args:
            data_dir: Directory containing data files. Defaults to ./data
        """
        self.data_dir = data_dir or Path('./data')
        self.data_dir.mkdir(parents=True, exist_ok=True)

    def read_csv(self, file_path: Union[str, Path]) -> pd.DataFrame:
        """
        Read application data from CSV file.

        Args:
            file_path: Path to the CSV file

        Returns:
            DataFrame containing application data

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns are missing
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Data file not found: {file_path}")

        try:
            df = pd.read_csv(file_path)
            logger.info(f"Successfully loaded {len(df)} applications from {file_path}")

            # Validate required columns
            missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
            if missing_cols:
                raise ValueError(f"Missing required columns: {missing_cols}")

            return df

        except Exception as e:
            logger.error(f"Error reading CSV file {file_path}: {e}")
            raise

    def read_excel(self, file_path: Union[str, Path], sheet_name: str = 'Applications') -> pd.DataFrame:
        """
        Read application data from Excel file.

        Args:
            file_path: Path to the Excel file
            sheet_name: Name of the sheet to read

        Returns:
            DataFrame containing application data

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns are missing
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Data file not found: {file_path}")

        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
            logger.info(f"Successfully loaded {len(df)} applications from {file_path}")

            # Validate required columns
            missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
            if missing_cols:
                raise ValueError(f"Missing required columns: {missing_cols}")

            return df

        except Exception as e:
            logger.error(f"Error reading Excel file {file_path}: {e}")
            raise

    def write_csv(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        include_timestamp: bool = True
    ) -> Path:
        """
        Write application data to CSV file.

        Args:
            df: DataFrame containing application data
            output_path: Path for the output file
            include_timestamp: Whether to append timestamp to filename

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Write to CSV
            df.to_csv(output_path, index=False)
            logger.info(f"Successfully wrote {len(df)} applications to {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error writing CSV file {output_path}: {e}")
            raise

    def write_excel(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        sheet_name: str = 'Applications',
        include_timestamp: bool = True
    ) -> Path:
        """
        Write application data to Excel file with formatting.

        Args:
            df: DataFrame containing application data
            output_path: Path for the output file
            sheet_name: Name of the sheet
            include_timestamp: Whether to append timestamp to filename

        Returns:
            Path to the written file
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            # Write to Excel with openpyxl engine for better formatting
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name=sheet_name, index=False)

                # Auto-adjust column widths
                worksheet = writer.sheets[sheet_name]
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width

            logger.info(f"Successfully wrote {len(df)} applications to {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error writing Excel file {output_path}: {e}")
            raise

    def validate_data(self, df: pd.DataFrame) -> tuple[bool, List[str]]:
        """
        Validate application data for completeness and correctness.

        Args:
            df: DataFrame to validate

        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        errors = []

        # Check for required columns
        missing_cols = set(self.REQUIRED_COLUMNS) - set(df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {missing_cols}")

        # Check for empty application names
        if 'Application Name' in df.columns:
            empty_names = df['Application Name'].isna().sum()
            if empty_names > 0:
                errors.append(f"{empty_names} applications have empty names")

        # Validate score ranges (0-10)
        score_columns = ['Business Value', 'Tech Health', 'Security', 'Strategic Fit']
        for col in score_columns:
            if col in df.columns:
                invalid = ((df[col] < 0) | (df[col] > 10)).sum()
                if invalid > 0:
                    errors.append(f"{invalid} invalid values in {col} (must be 0-10)")

        # Validate redundancy (0 or 1)
        if 'Redundancy' in df.columns:
            invalid = (~df['Redundancy'].isin([0, 1])).sum()
            if invalid > 0:
                errors.append(f"{invalid} invalid Redundancy values (must be 0 or 1)")

        # Validate cost (positive numbers)
        if 'Cost' in df.columns:
            invalid = (df['Cost'] < 0).sum()
            if invalid > 0:
                errors.append(f"{invalid} negative Cost values")

        # Validate usage (positive numbers)
        if 'Usage' in df.columns:
            invalid = (df['Usage'] < 0).sum()
            if invalid > 0:
                errors.append(f"{invalid} negative Usage values")

        is_valid = len(errors) == 0
        return is_valid, errors

    def get_summary_statistics(self, df: pd.DataFrame) -> Dict:
        """
        Calculate summary statistics for the application portfolio.

        Args:
            df: DataFrame containing application data

        Returns:
            Dictionary of summary statistics
        """
        stats = {
            'total_applications': len(df),
            'total_cost': df['Cost'].sum() if 'Cost' in df.columns else 0,
            'average_business_value': df['Business Value'].mean() if 'Business Value' in df.columns else 0,
            'average_tech_health': df['Tech Health'].mean() if 'Tech Health' in df.columns else 0,
            'average_security': df['Security'].mean() if 'Security' in df.columns else 0,
            'redundant_applications': df['Redundancy'].sum() if 'Redundancy' in df.columns else 0,
        }

        if 'Composite Score' in df.columns:
            stats['average_composite_score'] = df['Composite Score'].mean()
            stats['median_composite_score'] = df['Composite Score'].median()

        if 'Action Recommendation' in df.columns:
            stats['action_distribution'] = df['Action Recommendation'].value_counts().to_dict()

        return stats

    def filter_applications(
        self,
        df: pd.DataFrame,
        min_score: Optional[float] = None,
        max_score: Optional[float] = None,
        owner: Optional[str] = None,
        action: Optional[str] = None
    ) -> pd.DataFrame:
        """
        Filter applications based on criteria.

        Args:
            df: DataFrame to filter
            min_score: Minimum composite score
            max_score: Maximum composite score
            owner: Filter by owner
            action: Filter by action recommendation

        Returns:
            Filtered DataFrame
        """
        filtered_df = df.copy()

        if min_score is not None and 'Composite Score' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Composite Score'] >= min_score]

        if max_score is not None and 'Composite Score' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Composite Score'] <= max_score]

        if owner is not None and 'Owner' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Owner'].str.contains(owner, case=False, na=False)]

        if action is not None and 'Action Recommendation' in filtered_df.columns:
            filtered_df = filtered_df[filtered_df['Action Recommendation'] == action]

        return filtered_df

    def export_for_powerbi(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        include_timestamp: bool = True
    ) -> Path:
        """
        Export application data in Power BI-optimized Excel format.

        Creates a multi-sheet Excel workbook optimized for Power BI import with:
        - Properly formatted tables with headers
        - Normalized dimension scores table for relationships
        - TIME framework categorization table
        - Metadata and configuration information
        - Named ranges for easy Power BI data source configuration

        Args:
            df: DataFrame containing complete assessment data
            output_path: Path for the output Excel file
            include_timestamp: Whether to append timestamp to filename

        Returns:
            Path to the written Power BI-optimized Excel file

        Example:
            >>> handler = DataHandler()
            >>> path = handler.export_for_powerbi(results_df, 'powerbi_export.xlsx')
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            logger.info("Creating Power BI-optimized Excel export...")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Main Assessment Data (Fact Table)
                main_df = df.copy()

                # Add a unique ID column for relationships
                main_df.insert(0, 'Application_ID', range(1, len(main_df) + 1))

                # Ensure consistent data types
                if 'Cost' in main_df.columns:
                    main_df['Cost'] = pd.to_numeric(main_df['Cost'], errors='coerce')
                if 'Usage' in main_df.columns:
                    main_df['Usage'] = pd.to_numeric(main_df['Usage'], errors='coerce')

                main_df.to_excel(writer, sheet_name='Applications', index=False)

                # Sheet 2: Dimension Scores (normalized for Power BI relationships)
                dimension_cols = ['Business Value', 'Tech Health', 'Security',
                                'Strategic Fit', 'Usage', 'Cost', 'Redundancy']
                available_dims = [col for col in dimension_cols if col in df.columns]

                if available_dims:
                    # Create long-format dimension table
                    dimension_data = []
                    for idx, row in df.iterrows():
                        app_id = idx + 1
                        app_name = row.get('Application Name', f'App_{app_id}')
                        for dim in available_dims:
                            dimension_data.append({
                                'Application_ID': app_id,
                                'Application_Name': app_name,
                                'Dimension': dim,
                                'Score': row[dim]
                            })

                    dimension_df = pd.DataFrame(dimension_data)
                    dimension_df.to_excel(writer, sheet_name='Dimension_Scores', index=False)

                # Sheet 3: TIME Framework Data
                if 'TIME Category' in df.columns:
                    time_df = df[['Application Name', 'TIME Category']].copy()
                    time_df.insert(0, 'Application_ID', range(1, len(time_df) + 1))

                    # Add TIME scores if available
                    if 'TIME Business Value Score' in df.columns:
                        time_df['Business_Value_Score'] = df['TIME Business Value Score']
                    if 'TIME Technical Quality Score' in df.columns:
                        time_df['Technical_Quality_Score'] = df['TIME Technical Quality Score']
                    if 'TIME Rationale' in df.columns:
                        time_df['Rationale'] = df['TIME Rationale']

                    time_df.to_excel(writer, sheet_name='TIME_Framework', index=False)

                # Sheet 4: Action Recommendations
                if 'Action Recommendation' in df.columns:
                    action_df = df[['Application Name', 'Action Recommendation']].copy()
                    action_df.insert(0, 'Application_ID', range(1, len(action_df) + 1))

                    if 'Comments' in df.columns:
                        action_df['Comments'] = df['Comments']
                    if 'Composite Score' in df.columns:
                        action_df['Composite_Score'] = df['Composite Score']

                    action_df.to_excel(writer, sheet_name='Recommendations', index=False)

                # Sheet 5: Summary Statistics
                summary_data = {
                    'Metric': [
                        'Total Applications',
                        'Total Annual Cost',
                        'Average Business Value',
                        'Average Tech Health',
                        'Average Security',
                        'Average Composite Score',
                        'Redundant Applications'
                    ],
                    'Value': [
                        len(df),
                        df['Cost'].sum() if 'Cost' in df.columns else 0,
                        df['Business Value'].mean() if 'Business Value' in df.columns else 0,
                        df['Tech Health'].mean() if 'Tech Health' in df.columns else 0,
                        df['Security'].mean() if 'Security' in df.columns else 0,
                        df['Composite Score'].mean() if 'Composite Score' in df.columns else 0,
                        df['Redundancy'].sum() if 'Redundancy' in df.columns else 0
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='Summary_Stats', index=False)

                # Sheet 6: TIME Category Distribution
                if 'TIME Category' in df.columns:
                    time_dist = df['TIME Category'].value_counts().reset_index()
                    time_dist.columns = ['TIME_Category', 'Count']
                    time_dist['Percentage'] = (time_dist['Count'] / len(df) * 100).round(2)
                    time_dist.to_excel(writer, sheet_name='TIME_Distribution', index=False)

                # Sheet 7: Metadata
                metadata_data = {
                    'Property': [
                        'Export Date',
                        'Total Applications',
                        'Data Source',
                        'Export Version',
                        'Description'
                    ],
                    'Value': [
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                        len(df),
                        'Application Rationalization Assessment',
                        '1.0',
                        'Power BI optimized export with normalized tables and relationships'
                    ]
                }
                metadata_df = pd.DataFrame(metadata_data)
                metadata_df.to_excel(writer, sheet_name='Metadata', index=False)

            # Post-process: Add Excel tables for Power BI
            workbook = load_workbook(output_path)

            # Format as tables
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]
                if ws.max_row > 1:  # Has data beyond headers
                    # Get the data range
                    max_col_letter = ws.cell(1, ws.max_column).column_letter
                    table_ref = f"A1:{max_col_letter}{ws.max_row}"

                    # Create table
                    table_name = f"tbl{sheet_name.replace(' ', '_').replace('-', '_')}"
                    tab = Table(displayName=table_name, ref=table_ref)

                    # Add a default style
                    style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )
                    tab.tableStyleInfo = style

                    # Add table to worksheet
                    ws.add_table(tab)

                    # Format headers
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output_path)
            logger.info(f"Power BI export saved to: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error creating Power BI export {output_path}: {e}")
            raise

    def export_enhanced_excel(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        include_timestamp: bool = True,
        include_charts: bool = True
    ) -> Path:
        """
        Create enhanced Excel export with formatting, charts, and conditional formatting.

        This creates a professional, presentation-ready Excel workbook with:
        - Multiple worksheets (Summary, Detailed Scores, TIME Framework, Recommendations)
        - Conditional formatting (color-coded scores, priority highlighting)
        - Embedded charts (score distribution, TIME quadrant, category breakdown)
        - Professional styling (freeze panes, filters, formatted headers)
        - Summary dashboard worksheet with key metrics

        Args:
            df: DataFrame containing complete assessment data
            output_path: Path for the output Excel file
            include_timestamp: Whether to append timestamp to filename
            include_charts: Whether to include embedded charts

        Returns:
            Path to the written enhanced Excel file

        Example:
            >>> handler = DataHandler()
            >>> path = handler.export_enhanced_excel(results_df, 'assessment_report.xlsx')
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            logger.info("Creating enhanced Excel export with formatting and charts...")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Executive Summary Dashboard
                self._create_summary_dashboard(df, writer)

                # Sheet 2: Detailed Application Scores
                detailed_df = df.copy()
                detailed_df.to_excel(writer, sheet_name='Detailed_Scores', index=False)

                # Sheet 3: TIME Framework Analysis
                if 'TIME Category' in df.columns:
                    time_cols = ['Application Name', 'TIME Category']
                    if 'TIME Business Value Score' in df.columns:
                        time_cols.append('TIME Business Value Score')
                    if 'TIME Technical Quality Score' in df.columns:
                        time_cols.append('TIME Technical Quality Score')
                    if 'TIME Rationale' in df.columns:
                        time_cols.append('TIME Rationale')

                    time_df = df[time_cols].copy()
                    time_df.to_excel(writer, sheet_name='TIME_Framework', index=False)

                # Sheet 4: Recommendations
                if 'Action Recommendation' in df.columns:
                    rec_cols = ['Application Name', 'Action Recommendation', 'Composite Score']
                    if 'Comments' in df.columns:
                        rec_cols.append('Comments')

                    rec_df = df[[c for c in rec_cols if c in df.columns]].copy()
                    rec_df = rec_df.sort_values('Composite Score', ascending=False)
                    rec_df.to_excel(writer, sheet_name='Recommendations', index=False)

                # Sheet 5: Cost Analysis
                if 'Cost' in df.columns:
                    cost_cols = ['Application Name', 'Cost', 'Business Value', 'Composite Score']
                    cost_df = df[[c for c in cost_cols if c in df.columns]].copy()
                    cost_df = cost_df.sort_values('Cost', ascending=False)
                    cost_df.to_excel(writer, sheet_name='Cost_Analysis', index=False)

            # Post-process: Apply formatting and add charts
            workbook = load_workbook(output_path)

            # Format Summary Dashboard
            if 'Summary_Dashboard' in workbook.sheetnames:
                self._format_summary_dashboard(workbook['Summary_Dashboard'])

            # Format Detailed Scores sheet
            if 'Detailed_Scores' in workbook.sheetnames:
                self._format_detailed_scores(workbook['Detailed_Scores'], df)

            # Format TIME Framework sheet
            if 'TIME_Framework' in workbook.sheetnames:
                self._format_time_framework(workbook['TIME_Framework'], df, include_charts)

            # Format Recommendations sheet
            if 'Recommendations' in workbook.sheetnames:
                self._format_recommendations(workbook['Recommendations'], df, include_charts)

            # Format Cost Analysis sheet
            if 'Cost_Analysis' in workbook.sheetnames:
                self._format_cost_analysis(workbook['Cost_Analysis'], df, include_charts)

            workbook.save(output_path)
            logger.info(f"Enhanced Excel export saved to: {output_path}")

            return output_path

        except Exception as e:
            logger.error(f"Error creating enhanced Excel export {output_path}: {e}")
            raise

    def _create_summary_dashboard(self, df: pd.DataFrame, writer):
        """Create executive summary dashboard worksheet."""
        summary_data = []

        # Portfolio Overview
        summary_data.append(['PORTFOLIO OVERVIEW', ''])
        summary_data.append(['Total Applications', len(df)])
        summary_data.append([''])

        if 'Cost' in df.columns:
            summary_data.append(['Total Annual Cost', f"${df['Cost'].sum():,.0f}"])
            summary_data.append(['Average Cost per App', f"${df['Cost'].mean():,.0f}"])
            summary_data.append([''])

        # Average Scores
        summary_data.append(['AVERAGE SCORES', ''])
        score_metrics = ['Business Value', 'Tech Health', 'Security', 'Strategic Fit']
        for metric in score_metrics:
            if metric in df.columns:
                summary_data.append([f'Average {metric}', f"{df[metric].mean():.2f}/10"])

        if 'Composite Score' in df.columns:
            summary_data.append(['Average Composite Score', f"{df['Composite Score'].mean():.2f}/100"])

        summary_data.append([''])

        # TIME Framework Distribution
        if 'TIME Category' in df.columns:
            summary_data.append(['TIME FRAMEWORK DISTRIBUTION', ''])
            time_counts = df['TIME Category'].value_counts()
            for category, count in time_counts.items():
                percentage = (count / len(df)) * 100
                summary_data.append([category, f"{count} ({percentage:.1f}%)"])

        summary_data.append([''])

        # Action Recommendations
        if 'Action Recommendation' in df.columns:
            summary_data.append(['ACTION RECOMMENDATIONS', ''])
            action_counts = df['Action Recommendation'].value_counts().head(5)
            for action, count in action_counts.items():
                percentage = (count / len(df)) * 100
                summary_data.append([action, f"{count} ({percentage:.1f}%)"])

        summary_data.append([''])

        # Key Risks
        summary_data.append(['KEY METRICS', ''])
        if 'Redundancy' in df.columns:
            redundant_count = df['Redundancy'].sum()
            summary_data.append(['Redundant Applications', int(redundant_count)])

        if 'Security' in df.columns:
            low_security_count = len(df[df['Security'] < 5])
            summary_data.append(['Low Security Score (<5)', low_security_count])

        if 'Tech Health' in df.columns:
            poor_health_count = len(df[df['Tech Health'] < 5])
            summary_data.append(['Poor Tech Health (<5)', poor_health_count])

        summary_df = pd.DataFrame(summary_data, columns=['Metric', 'Value'])
        summary_df.to_excel(writer, sheet_name='Summary_Dashboard', index=False, header=False)

    def _format_summary_dashboard(self, ws):
        """Apply formatting to summary dashboard."""
        # Header font for section titles
        header_font = Font(bold=True, size=12, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Value font
        value_font = Font(size=11)

        # Format cells
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            cell_a = row[0]
            cell_b = row[1] if len(row) > 1 else None

            # Check if this is a section header
            if cell_a.value and isinstance(cell_a.value, str) and cell_a.value.isupper():
                cell_a.font = header_font
                cell_a.fill = header_fill
                if cell_b:
                    cell_b.font = header_font
                    cell_b.fill = header_fill
            else:
                cell_a.font = Font(bold=True, size=10)
                if cell_b:
                    cell_b.font = value_font

            cell_a.alignment = Alignment(horizontal="left", vertical="center")
            if cell_b:
                cell_b.alignment = Alignment(horizontal="right", vertical="center")

        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 25

        # Freeze top row
        ws.freeze_panes = 'A2'

    def _format_detailed_scores(self, ws, df):
        """Apply conditional formatting to detailed scores sheet."""
        # Header row formatting
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Apply conditional formatting for score columns
        score_columns = {
            'Business Value': None,
            'Tech Health': None,
            'Security': None,
            'Strategic Fit': None,
            'Composite Score': None
        }

        # Find column indices
        for idx, cell in enumerate(ws[1], 1):
            if cell.value in score_columns:
                score_columns[cell.value] = idx

        # Apply color scale formatting
        for col_name, col_idx in score_columns.items():
            if col_idx:
                col_letter = ws.cell(1, col_idx).column_letter
                if col_name == 'Composite Score':
                    # 0-100 scale
                    ws.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{ws.max_row}',
                        ColorScaleRule(
                            start_type='num', start_value=0, start_color='F8696B',
                            mid_type='num', mid_value=50, mid_color='FFEB84',
                            end_type='num', end_value=100, end_color='63BE7B'
                        )
                    )
                else:
                    # 0-10 scale
                    ws.conditional_formatting.add(
                        f'{col_letter}2:{col_letter}{ws.max_row}',
                        ColorScaleRule(
                            start_type='num', start_value=0, start_color='F8696B',
                            mid_type='num', mid_value=5, mid_color='FFEB84',
                            end_type='num', end_value=10, end_color='63BE7B'
                        )
                    )

        # Freeze panes
        ws.freeze_panes = 'B2'

        # Enable auto-filter
        ws.auto_filter.ref = ws.dimensions

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    def _format_time_framework(self, ws, df, include_charts: bool):
        """Format TIME Framework sheet with conditional formatting and optional charts."""
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Find TIME Category column
        category_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'TIME Category':
                category_col = idx
                break

        # Apply color coding to TIME categories
        if category_col:
            col_letter = ws.cell(1, category_col).column_letter
            time_colors = {
                'Invest': '2E7D32',
                'Tolerate': 'FFA726',
                'Migrate': '1976D2',
                'Eliminate': 'C62828'
            }

            for category, color in time_colors.items():
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{ws.max_row}',
                    CellIsRule(
                        operator='equal',
                        formula=[f'"{category}"'],
                        fill=PatternFill(start_color=color, end_color=color, fill_type="solid"),
                        font=Font(color="FFFFFF", bold=True)
                    )
                )

        # Add pie chart for TIME distribution
        if include_charts and 'TIME Category' in df.columns and ws.max_row > 1:
            # Count categories
            time_counts = df['TIME Category'].value_counts()

            # Create chart data area (to the right of the main data)
            chart_start_col = ws.max_column + 2
            ws.cell(1, chart_start_col, 'Category')
            ws.cell(1, chart_start_col + 1, 'Count')

            for idx, (category, count) in enumerate(time_counts.items(), 2):
                ws.cell(idx, chart_start_col, category)
                ws.cell(idx, chart_start_col + 1, count)

            # Create pie chart
            pie = PieChart()
            labels = Reference(ws, min_col=chart_start_col, min_row=2, max_row=len(time_counts) + 1)
            data = Reference(ws, min_col=chart_start_col + 1, min_row=1, max_row=len(time_counts) + 1)
            pie.add_data(data, titles_from_data=True)
            pie.set_categories(labels)
            pie.title = "TIME Framework Distribution"
            pie.height = 10
            pie.width = 15

            ws.add_chart(pie, f"{ws.cell(5, chart_start_col + 3).coordinate}")

        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = f'A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}'

    def _format_recommendations(self, ws, df, include_charts: bool):
        """Format Recommendations sheet."""
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Find Composite Score column for conditional formatting
        score_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Composite Score':
                score_col = idx
                break

        if score_col:
            col_letter = ws.cell(1, score_col).column_letter
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='num', start_value=0, start_color='F8696B',
                    mid_type='num', mid_value=50, mid_color='FFEB84',
                    end_type='num', end_value=100, end_color='63BE7B'
                )
            )

        # Add bar chart for recommendations
        if include_charts and 'Action Recommendation' in df.columns and ws.max_row > 1:
            # Count recommendations
            rec_counts = df['Action Recommendation'].value_counts().head(5)

            # Create chart data area
            chart_start_col = ws.max_column + 2
            ws.cell(1, chart_start_col, 'Action')
            ws.cell(1, chart_start_col + 1, 'Count')

            for idx, (action, count) in enumerate(rec_counts.items(), 2):
                action_short = action[:30] + '...' if len(action) > 30 else action
                ws.cell(idx, chart_start_col, action_short)
                ws.cell(idx, chart_start_col + 1, count)

            # Create bar chart
            chart = BarChart()
            chart.type = "col"
            chart.title = "Top Action Recommendations"
            chart.y_axis.title = 'Count'
            chart.x_axis.title = 'Action'

            data = Reference(ws, min_col=chart_start_col + 1, min_row=1, max_row=min(len(rec_counts) + 1, 6))
            cats = Reference(ws, min_col=chart_start_col, min_row=2, max_row=min(len(rec_counts) + 1, 6))
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 15

            ws.add_chart(chart, f"{ws.cell(5, chart_start_col + 3).coordinate}")

        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = f'A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}'

    def _format_cost_analysis(self, ws, df, include_charts: bool):
        """Format Cost Analysis sheet."""
        # Header formatting
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Find Cost column
        cost_col = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == 'Cost':
                cost_col = idx
                # Format as currency
                col_letter = ws.cell(1, cost_col).column_letter
                for row in range(2, ws.max_row + 1):
                    ws[f'{col_letter}{row}'].number_format = '$#,##0'
                break

        # Apply conditional formatting to cost (higher cost = redder)
        if cost_col and ws.max_row > 1:
            col_letter = ws.cell(1, cost_col).column_letter
            ws.conditional_formatting.add(
                f'{col_letter}2:{col_letter}{ws.max_row}',
                ColorScaleRule(
                    start_type='min', start_color='63BE7B',
                    mid_type='percentile', mid_value=50, mid_color='FFEB84',
                    end_type='max', end_color='F8696B'
                )
            )

        ws.freeze_panes = 'B2'
        ws.auto_filter.ref = f'A1:{ws.cell(1, ws.max_column).column_letter}{ws.max_row}'

    def export_for_tableau(
        self,
        df: pd.DataFrame,
        output_path: Union[str, Path],
        include_timestamp: bool = True
    ) -> Path:
        """
        Export application data in Tableau-optimized format.

        Creates a clean, denormalized CSV file optimized for Tableau with:
        - Flat structure (no hierarchies)
        - Clean column names (no special characters)
        - Proper data types
        - Date fields for time-based analysis
        - Calculated fields pre-computed

        Args:
            df: DataFrame containing complete assessment data
            output_path: Path for the output CSV file
            include_timestamp: Whether to append timestamp to filename

        Returns:
            Path to the written Tableau-optimized CSV file

        Example:
            >>> handler = DataHandler()
            >>> path = handler.export_for_tableau(results_df, 'tableau_data.csv')
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            # Ensure output directory exists
            output_path.parent.mkdir(parents=True, exist_ok=True)

            logger.info("Creating Tableau-optimized CSV export...")

            # Create Tableau-friendly DataFrame
            tableau_df = df.copy()

            # Add assessment date for time-based analysis
            tableau_df.insert(0, 'Assessment_Date', datetime.now().strftime('%Y-%m-%d'))

            # Clean column names (remove spaces, special characters)
            tableau_df.columns = [
                col.replace(' ', '_').replace('(', '').replace(')', '').replace('/', '_')
                for col in tableau_df.columns
            ]

            # Add calculated fields useful for Tableau
            if 'Composite_Score' in tableau_df.columns:
                # Score category
                tableau_df['Score_Category'] = pd.cut(
                    tableau_df['Composite_Score'],
                    bins=[0, 30, 50, 70, 100],
                    labels=['Poor', 'Fair', 'Good', 'Excellent']
                )

                # Performance tier
                tableau_df['Performance_Tier'] = pd.qcut(
                    tableau_df['Composite_Score'],
                    q=4,
                    labels=['Bottom 25%', '25-50%', '50-75%', 'Top 25%'],
                    duplicates='drop'
                )

            # Add risk flags
            if 'Security' in tableau_df.columns and 'Business_Value' in tableau_df.columns:
                tableau_df['High_Risk_Flag'] = (
                    (tableau_df['Security'] < 5) & (tableau_df['Business_Value'] > 7)
                ).astype(int)

            # Add cost efficiency metric
            if 'Cost' in tableau_df.columns and 'Business_Value' in tableau_df.columns:
                # Avoid division by zero
                cost_nonzero = tableau_df['Cost'].replace(0, 1)
                tableau_df['Value_per_Dollar'] = (
                    tableau_df['Business_Value'] / cost_nonzero * 10000
                ).round(2)

            # Ensure consistent data types
            # Convert any remaining NaN to None for better Tableau handling
            tableau_df = tableau_df.replace({np.nan: None})

            # Write to CSV with UTF-8 encoding (Tableau standard)
            tableau_df.to_csv(output_path, index=False, encoding='utf-8')

            logger.info(f"Tableau export saved to: {output_path}")
            logger.info(f"Exported {len(tableau_df)} rows with {len(tableau_df.columns)} columns")

            return output_path

        except Exception as e:
            logger.error(f"Error creating Tableau export {output_path}: {e}")
            raise

    def generate_complete_report_bundle(
        self,
        df: pd.DataFrame,
        output_dir: Union[str, Path],
        report_name: str = 'assessment_report',
        include_visualizations: bool = False
    ) -> Dict[str, Path]:
        """
        Generate a complete report bundle with all export formats.

        Creates a comprehensive report package including:
        - CSV data export
        - Power BI-optimized Excel
        - Enhanced Excel with charts
        - Tableau-optimized CSV
        - README with usage instructions
        - Optionally: visualizations

        Args:
            df: DataFrame containing complete assessment data
            output_dir: Directory for the report bundle
            report_name: Base name for the report files
            include_visualizations: Whether to generate visualizations

        Returns:
            Dictionary mapping export type to file path

        Example:
            >>> handler = DataHandler()
            >>> bundle = handler.generate_complete_report_bundle(
            ...     results_df,
            ...     'output/reports/2025-Q1',
            ...     report_name='Q1_Assessment'
            ... )
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        logger.info(f"Generating complete report bundle in: {output_dir}")

        bundle_files = {}
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')

        try:
            # 1. CSV Export
            logger.info("Creating CSV export...")
            csv_path = self.write_csv(
                df,
                output_dir / f'{report_name}_data.csv',
                include_timestamp=False
            )
            bundle_files['csv'] = csv_path

            # 2. Power BI Export
            logger.info("Creating Power BI export...")
            powerbi_path = self.export_for_powerbi(
                df,
                output_dir / f'{report_name}_powerbi.xlsx',
                include_timestamp=False
            )
            bundle_files['powerbi'] = powerbi_path

            # 3. Enhanced Excel Report
            logger.info("Creating enhanced Excel report...")
            excel_path = self.export_enhanced_excel(
                df,
                output_dir / f'{report_name}_executive.xlsx',
                include_timestamp=False,
                include_charts=True
            )
            bundle_files['excel'] = excel_path

            # 4. Tableau Export
            logger.info("Creating Tableau export...")
            tableau_path = self.export_for_tableau(
                df,
                output_dir / f'{report_name}_tableau.csv',
                include_timestamp=False
            )
            bundle_files['tableau'] = tableau_path

            # 5. Generate visualizations if requested
            if include_visualizations:
                logger.info("Generating visualizations...")
                viz_dir = output_dir / 'visualizations'
                viz_dir.mkdir(exist_ok=True)

                from .visualizations import VisualizationEngine

                viz_engine = VisualizationEngine(output_dir=viz_dir)

                try:
                    bundle_files['viz_heatmap'] = viz_engine.create_score_heatmap(
                        df, output_file='score_heatmap.png'
                    )
                    bundle_files['viz_time_quadrant'] = viz_engine.create_time_quadrant_heatmap(
                        df, output_file='time_quadrant.png'
                    )
                    bundle_files['viz_dashboard'] = viz_engine.create_comprehensive_dashboard(
                        df, output_file='dashboard.png'
                    )
                except Exception as e:
                    logger.warning(f"Some visualizations failed: {e}")

            # 6. Create README with instructions
            logger.info("Creating README...")
            readme_path = self._create_bundle_readme(output_dir, report_name, bundle_files, df)
            bundle_files['readme'] = readme_path

            # 7. Create summary statistics file
            logger.info("Creating summary statistics...")
            summary_path = self._create_summary_stats(output_dir, report_name, df)
            bundle_files['summary'] = summary_path

            logger.info(f"Report bundle complete! Generated {len(bundle_files)} files")

            return bundle_files

        except Exception as e:
            logger.error(f"Error generating report bundle: {e}")
            raise

    def _create_bundle_readme(
        self,
        output_dir: Path,
        report_name: str,
        files: Dict[str, Path],
        df: pd.DataFrame
    ) -> Path:
        """Create README file for report bundle."""
        readme_path = output_dir / 'README.md'

        stats = self.get_summary_statistics(df)

        content = f"""# Application Rationalization Assessment Report
## {report_name}

**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
**Applications Assessed:** {len(df)}
**Total Annual Cost:** ${stats.get('total_cost', 0):,.0f}
**Average Composite Score:** {stats.get('average_composite_score', 0):.1f}/100

---

## Report Contents

This report bundle contains assessment data in multiple formats for different use cases:

### 1. CSV Data Export
**File:** `{report_name}_data.csv`
**Use for:** General data analysis, archival, version control
**Format:** Standard CSV with all assessment data

### 2. Power BI Export
**File:** `{report_name}_powerbi.xlsx`
**Use for:** Interactive dashboards, drill-down analysis
**Format:** Multi-sheet Excel optimized for Power BI

**Import Instructions:**
1. Open Power BI Desktop
2. Get Data → Excel → Select `{report_name}_powerbi.xlsx`
3. Load these tables: Applications, Dimension_Scores, TIME_Framework
4. Create relationships using Application_ID field
5. Build visualizations

### 3. Enhanced Excel Report
**File:** `{report_name}_executive.xlsx`
**Use for:** Executive presentations, board reports
**Format:** Formatted Excel with charts and conditional formatting

**Features:**
- Summary dashboard with key metrics
- Detailed scores with color coding
- TIME framework analysis with charts
- Action recommendations
- Cost analysis

### 4. Tableau Export
**File:** `{report_name}_tableau.csv`
**Use for:** Tableau dashboards and visualizations
**Format:** Clean CSV with calculated fields

**Import Instructions:**
1. Open Tableau Desktop
2. Connect to Data → Text file → Select `{report_name}_tableau.csv`
3. Suggested visualizations:
   - Scatter plot: Business_Value vs Tech_Health (colored by TIME_Category)
   - Bar chart: Top applications by Composite_Score
   - Heatmap: Score_Category distribution
   - Treemap: Applications sized by Cost

"""

        if 'viz_dashboard' in files:
            content += """
### 5. Visualizations
**Folder:** `visualizations/`
**Files:**
- `score_heatmap.png` - Application score matrix
- `time_quadrant.png` - TIME framework positioning
- `dashboard.png` - Comprehensive executive dashboard

"""

        content += f"""
---

## Summary Statistics

- **Total Applications:** {stats['total_applications']}
- **Total Annual Cost:** ${stats['total_cost']:,.0f}
- **Average Business Value:** {stats['average_business_value']:.2f}/10
- **Average Tech Health:** {stats['average_tech_health']:.2f}/10
- **Average Security:** {stats['average_security']:.2f}/10
- **Redundant Applications:** {int(stats['redundant_applications'])}

"""

        if 'action_distribution' in stats:
            content += "\n### Action Recommendations:\n"
            for action, count in stats['action_distribution'].items():
                content += f"- **{action}:** {count} applications\n"

        content += """
---

## Next Steps

1. **Review Executive Report** - Start with the enhanced Excel file for high-level insights
2. **Import to BI Tool** - Use Power BI or Tableau exports for interactive analysis
3. **Share with Stakeholders** - Distribute appropriate formats to different audiences
4. **Plan Actions** - Focus on applications flagged for immediate action
5. **Track Over Time** - Compare with future assessments to measure progress

## Support

For questions about this assessment or the tool:
- Review full documentation in the repository
- Check the visualization guide for chart interpretations
- See TIME framework documentation for category definitions

---

*Generated by Application Rationalization Assessment Tool*
"""

        readme_path.write_text(content)
        return readme_path

    def _create_summary_stats(self, output_dir: Path, report_name: str, df: pd.DataFrame) -> Path:
        """Create summary statistics file."""
        summary_path = output_dir / f'{report_name}_summary.txt'

        stats = self.get_summary_statistics(df)

        content = f"""APPLICATION RATIONALIZATION ASSESSMENT - SUMMARY STATISTICS
{'=' * 70}

Report: {report_name}
Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}

PORTFOLIO OVERVIEW
{'=' * 70}
Total Applications: {stats['total_applications']}
Total Annual Cost: ${stats['total_cost']:,.0f}
Average Cost per App: ${stats['total_cost'] / stats['total_applications']:,.0f}

AVERAGE SCORES (0-10 scale)
{'=' * 70}
Business Value:  {stats['average_business_value']:.2f}
Tech Health:     {stats['average_tech_health']:.2f}
Security:        {stats['average_security']:.2f}
"""

        if 'average_composite_score' in stats:
            content += f"\nCOMPOSITE SCORES (0-100 scale)\n{'=' * 70}\n"
            content += f"Average: {stats['average_composite_score']:.2f}\n"
            content += f"Median:  {stats['median_composite_score']:.2f}\n"

        if 'action_distribution' in stats:
            content += f"\nACTION RECOMMENDATIONS\n{'=' * 70}\n"
            for action, count in sorted(stats['action_distribution'].items(), key=lambda x: x[1], reverse=True):
                percentage = (count / stats['total_applications']) * 100
                content += f"{action:30s} {count:3d} ({percentage:5.1f}%)\n"

        content += f"\nKEY METRICS\n{'=' * 70}\n"
        content += f"Redundant Applications: {int(stats['redundant_applications'])}\n"

        if 'Security' in df.columns:
            low_security = len(df[df['Security'] < 5])
            content += f"Low Security Score (<5): {low_security}\n"

        if 'Tech Health' in df.columns:
            poor_health = len(df[df['Tech Health'] < 5])
            content += f"Poor Tech Health (<5):  {poor_health}\n"

        summary_path.write_text(content)
        return summary_path

    # ==================================================================================
    # STAKEHOLDER SURVEY INTEGRATION METHODS
    # ==================================================================================

    SURVEY_REQUIRED_COLUMNS = [
        'Application Name',
        'Stakeholder Name',
        'Stakeholder Role',
        'Survey Date'
    ]

    SURVEY_RATING_COLUMNS = [
        'Critical to Business',
        'Easy to Replace',
        'User Satisfaction',
        'Performance Rating',
        'Strategic Importance'
    ]

    def read_survey_data(self, file_path: Union[str, Path]) -> pd.DataFrame:
        """
        Read stakeholder survey data from CSV file.

        Survey data should include stakeholder feedback on applications with
        ratings and qualitative comments.

        Args:
            file_path: Path to the survey CSV file

        Returns:
            DataFrame containing survey data

        Raises:
            FileNotFoundError: If file doesn't exist
            ValueError: If required columns are missing

        Example:
            >>> handler = DataHandler()
            >>> survey_df = handler.read_survey_data('data/sample_survey.csv')
        """
        file_path = Path(file_path)

        if not file_path.exists():
            raise FileNotFoundError(f"Survey file not found: {file_path}")

        try:
            df = pd.read_csv(file_path)
            logger.info(f"Successfully loaded {len(df)} survey responses from {file_path}")

            # Validate required columns
            missing_cols = set(self.SURVEY_REQUIRED_COLUMNS) - set(df.columns)
            if missing_cols:
                raise ValueError(f"Missing required survey columns: {missing_cols}")

            # Parse survey dates
            if 'Survey Date' in df.columns:
                df['Survey Date'] = pd.to_datetime(df['Survey Date'], errors='coerce')

            # Ensure rating columns are numeric
            for col in self.SURVEY_RATING_COLUMNS:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')

            return df

        except Exception as e:
            logger.error(f"Error reading survey file {file_path}: {e}")
            raise

    def validate_survey_data(self, survey_df: pd.DataFrame) -> tuple[bool, List[str]]:
        """
        Validate survey data for completeness and correctness.

        Args:
            survey_df: DataFrame containing survey data

        Returns:
            Tuple of (is_valid, list_of_errors)

        Example:
            >>> handler = DataHandler()
            >>> survey_df = handler.read_survey_data('data/sample_survey.csv')
            >>> is_valid, errors = handler.validate_survey_data(survey_df)
        """
        errors = []

        # Check for required columns
        missing_cols = set(self.SURVEY_REQUIRED_COLUMNS) - set(survey_df.columns)
        if missing_cols:
            errors.append(f"Missing required columns: {missing_cols}")

        # Check for empty application names
        if 'Application Name' in survey_df.columns:
            empty_apps = survey_df['Application Name'].isna().sum()
            if empty_apps > 0:
                errors.append(f"{empty_apps} survey responses have empty application names")

        # Check for empty stakeholder names
        if 'Stakeholder Name' in survey_df.columns:
            empty_stakeholders = survey_df['Stakeholder Name'].isna().sum()
            if empty_stakeholders > 0:
                errors.append(f"{empty_stakeholders} survey responses have empty stakeholder names")

        # Validate rating ranges (1-5 scale)
        for col in self.SURVEY_RATING_COLUMNS:
            if col in survey_df.columns:
                valid_ratings = survey_df[col].dropna()
                invalid = ((valid_ratings < 1) | (valid_ratings > 5)).sum()
                if invalid > 0:
                    errors.append(f"{invalid} invalid values in {col} (must be 1-5)")

        # Check for invalid dates
        if 'Survey Date' in survey_df.columns:
            invalid_dates = survey_df['Survey Date'].isna().sum()
            if invalid_dates > 0:
                errors.append(f"{invalid_dates} survey responses have invalid dates")

        is_valid = len(errors) == 0
        return is_valid, errors

    def aggregate_survey_responses(
        self,
        survey_df: pd.DataFrame,
        method: str = 'mean'
    ) -> pd.DataFrame:
        """
        Aggregate multiple stakeholder responses per application.

        When multiple stakeholders provide feedback for the same application,
        this method aggregates their responses into single scores per application.

        Args:
            survey_df: DataFrame containing raw survey data
            method: Method to aggregate ('mean', 'median', 'weighted')

        Returns:
            DataFrame with one row per application containing aggregated scores

        Example:
            >>> handler = DataHandler()
            >>> survey_df = handler.read_survey_data('data/sample_survey.csv')
            >>> agg_df = handler.aggregate_survey_responses(survey_df, 'mean')
        """
        if method not in ['mean', 'median', 'weighted']:
            raise ValueError(f"Invalid aggregation method: {method}")

        # Group by application
        grouped = survey_df.groupby('Application Name')

        # Determine aggregation function
        agg_func = np.mean if method == 'mean' else np.median

        # Aggregate numeric columns
        agg_dict = {}
        for col in self.SURVEY_RATING_COLUMNS:
            if col in survey_df.columns:
                agg_dict[col] = agg_func

        # Add response count
        agg_dict['Response Count'] = 'count'

        # Perform aggregation
        if method == 'mean':
            aggregated = grouped.agg({
                **{col: 'mean' for col in self.SURVEY_RATING_COLUMNS if col in survey_df.columns},
                'Stakeholder Name': 'count',
                'Qualitative Feedback': lambda x: ' | '.join(x.dropna().astype(str))
            })
        elif method == 'median':
            aggregated = grouped.agg({
                **{col: 'median' for col in self.SURVEY_RATING_COLUMNS if col in survey_df.columns},
                'Stakeholder Name': 'count',
                'Qualitative Feedback': lambda x: ' | '.join(x.dropna().astype(str))
            })
        else:  # weighted - weight by stakeholder seniority (simplified: all equal for now)
            aggregated = grouped.agg({
                **{col: 'mean' for col in self.SURVEY_RATING_COLUMNS if col in survey_df.columns},
                'Stakeholder Name': 'count',
                'Qualitative Feedback': lambda x: ' | '.join(x.dropna().astype(str))
            })

        # Rename count column
        aggregated = aggregated.rename(columns={'Stakeholder Name': 'Survey Response Count'})

        # Calculate consensus metrics (standard deviation to measure agreement)
        consensus_metrics = {}
        for col in self.SURVEY_RATING_COLUMNS:
            if col in survey_df.columns:
                consensus_metrics[f'{col} Consensus'] = grouped[col].std()

        consensus_df = pd.DataFrame(consensus_metrics)

        # Merge aggregated scores with consensus metrics
        result = aggregated.join(consensus_df)

        # Calculate overall consensus score (lower std = higher consensus)
        consensus_cols = [col for col in result.columns if 'Consensus' in col]
        if consensus_cols:
            result['Overall Consensus Score'] = (
                5 - result[consensus_cols].mean(axis=1)
            ).clip(1, 5)  # Invert so high score = high consensus

        result = result.reset_index()

        logger.info(f"Aggregated {len(survey_df)} responses into {len(result)} applications")

        return result

    def merge_survey_with_assessment(
        self,
        assessment_df: pd.DataFrame,
        survey_df: pd.DataFrame,
        survey_weight: float = 0.3
    ) -> pd.DataFrame:
        """
        Merge stakeholder survey data with quantitative assessment scores.

        This method combines survey feedback with existing assessment scores,
        creating adjusted scores that reflect both quantitative metrics and
        stakeholder sentiment.

        Args:
            assessment_df: DataFrame with quantitative assessment data
            survey_df: DataFrame with aggregated survey data
            survey_weight: Weight given to survey data (0-1), default 0.3 (30%)

        Returns:
            DataFrame with merged assessment and survey data, including:
            - Original quantitative scores
            - Survey ratings
            - Survey-adjusted scores
            - Variance analysis (differences between quantitative and qualitative)

        Example:
            >>> handler = DataHandler()
            >>> assessment_df = handler.read_csv('results.csv')
            >>> survey_df = handler.read_survey_data('survey.csv')
            >>> agg_survey = handler.aggregate_survey_responses(survey_df)
            >>> merged = handler.merge_survey_with_assessment(assessment_df, agg_survey, 0.3)
        """
        if not 0 <= survey_weight <= 1:
            raise ValueError("survey_weight must be between 0 and 1")

        # Merge on application name
        merged = assessment_df.merge(
            survey_df,
            on='Application Name',
            how='left',
            suffixes=('', '_Survey')
        )

        # Store original scores
        if 'Business Value' in merged.columns:
            merged['Business Value Original'] = merged['Business Value']

        # Map survey ratings to assessment scores (1-5 scale to 0-10 scale)
        survey_to_assessment_mapping = {
            'Critical to Business': 'Business Value',
            'User Satisfaction': 'Usage',
            'Performance Rating': 'Tech Health',
            'Strategic Importance': 'Strategic Fit'
        }

        # Calculate survey-adjusted scores
        for survey_col, assessment_col in survey_to_assessment_mapping.items():
            if survey_col in merged.columns and assessment_col in merged.columns:
                # Convert 1-5 scale to 0-10 scale
                survey_score_scaled = (merged[survey_col] - 1) * 2.5

                # Store original
                merged[f'{assessment_col} Original'] = merged[assessment_col]

                # Calculate adjusted score (weighted average)
                merged[f'{assessment_col} Survey Adjusted'] = (
                    merged[assessment_col] * (1 - survey_weight) +
                    survey_score_scaled * survey_weight
                ).round(2)

                # Calculate variance (difference between quantitative and qualitative)
                merged[f'{assessment_col} Variance'] = (
                    survey_score_scaled - merged[assessment_col]
                ).round(2)

        # Calculate "Easy to Replace" inverse score (lower = more critical)
        if 'Easy to Replace' in merged.columns:
            merged['Replacement Risk'] = (6 - merged['Easy to Replace']).clip(1, 5)

        # Add survey metadata
        merged['Has Survey Data'] = merged['Survey Response Count'].notna()
        merged['Survey Response Count'] = merged['Survey Response Count'].fillna(0).astype(int)

        # Calculate high variance flags (significant difference between scores)
        variance_cols = [col for col in merged.columns if 'Variance' in col]
        if variance_cols:
            merged['High Variance Flag'] = (
                merged[variance_cols].abs().max(axis=1) > 2
            ).astype(int)

        logger.info(f"Merged {len(merged)} applications with survey data")
        logger.info(f"{merged['Has Survey Data'].sum()} applications have survey responses")

        return merged

    def calculate_survey_impact(
        self,
        merged_df: pd.DataFrame
    ) -> Dict:
        """
        Analyze the impact of survey data on assessment scores.

        Calculates statistics showing how stakeholder feedback differs from
        quantitative scores and identifies applications with significant variances.

        Args:
            merged_df: DataFrame from merge_survey_with_assessment()

        Returns:
            Dictionary containing:
            - variance_summary: Statistics on score variances
            - high_variance_apps: Applications with significant differences
            - consensus_summary: Stakeholder agreement metrics
            - sentiment_analysis: Overall stakeholder sentiment by app

        Example:
            >>> handler = DataHandler()
            >>> impact = handler.calculate_survey_impact(merged_df)
            >>> print(impact['variance_summary'])
        """
        impact = {}

        # Variance analysis
        variance_cols = [col for col in merged_df.columns if 'Variance' in col and 'Flag' not in col]

        if variance_cols:
            variance_summary = {}
            for col in variance_cols:
                variance_summary[col] = {
                    'mean': merged_df[col].mean(),
                    'median': merged_df[col].median(),
                    'std': merged_df[col].std(),
                    'max_positive': merged_df[col].max(),
                    'max_negative': merged_df[col].min()
                }
            impact['variance_summary'] = variance_summary

        # High variance applications
        if 'High Variance Flag' in merged_df.columns:
            high_var_apps = merged_df[merged_df['High Variance Flag'] == 1][[
                'Application Name',
                *[col for col in variance_cols],
                'Survey Response Count'
            ]]
            impact['high_variance_apps'] = high_var_apps.to_dict('records')

        # Consensus analysis
        if 'Overall Consensus Score' in merged_df.columns:
            impact['consensus_summary'] = {
                'average_consensus': merged_df['Overall Consensus Score'].mean(),
                'high_consensus_count': len(merged_df[merged_df['Overall Consensus Score'] >= 4]),
                'low_consensus_count': len(merged_df[merged_df['Overall Consensus Score'] < 3])
            }

        # Sentiment analysis (based on survey ratings)
        if 'Critical to Business' in merged_df.columns and 'User Satisfaction' in merged_df.columns:
            merged_df_with_survey = merged_df[merged_df['Has Survey Data'] == True]

            sentiment_categories = []
            for _, row in merged_df_with_survey.iterrows():
                critical = row.get('Critical to Business', 0)
                satisfaction = row.get('User Satisfaction', 0)

                if critical >= 4 and satisfaction >= 4:
                    category = 'High Value & Satisfaction'
                elif critical >= 4 and satisfaction < 3:
                    category = 'High Value but Poor Satisfaction'
                elif critical < 3 and satisfaction >= 4:
                    category = 'Low Value but High Satisfaction'
                else:
                    category = 'Low Value & Satisfaction'

                sentiment_categories.append(category)

            sentiment_dist = pd.Series(sentiment_categories).value_counts().to_dict()
            impact['sentiment_analysis'] = sentiment_dist

        # Applications needing attention (high criticality, low satisfaction)
        if 'Critical to Business' in merged_df.columns and 'User Satisfaction' in merged_df.columns:
            needs_attention = merged_df[
                (merged_df['Critical to Business'] >= 4) &
                (merged_df['User Satisfaction'] < 3)
            ][['Application Name', 'Critical to Business', 'User Satisfaction', 'Qualitative Feedback']]

            impact['needs_attention'] = needs_attention.to_dict('records')

        logger.info("Survey impact analysis complete")

        return impact

    def export_survey_analysis(
        self,
        merged_df: pd.DataFrame,
        output_path: Union[str, Path],
        include_timestamp: bool = True
    ) -> Path:
        """
        Export survey analysis results to Excel with multiple worksheets.

        Creates a comprehensive Excel workbook with:
        - Merged data with survey-adjusted scores
        - Variance analysis
        - Consensus metrics
        - Qualitative feedback summary

        Args:
            merged_df: DataFrame from merge_survey_with_assessment()
            output_path: Path for output Excel file
            include_timestamp: Whether to include timestamp in filename

        Returns:
            Path to exported file

        Example:
            >>> handler = DataHandler()
            >>> path = handler.export_survey_analysis(merged_df, 'survey_analysis.xlsx')
        """
        output_path = Path(output_path)

        if include_timestamp:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            output_path = output_path.parent / f"{output_path.stem}_{timestamp}{output_path.suffix}"

        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)

            logger.info("Creating survey analysis Excel export...")

            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # Sheet 1: Merged data with all scores
                main_cols = [
                    'Application Name',
                    'Business Value Original',
                    'Business Value Survey Adjusted',
                    'Business Value Variance',
                    'Tech Health Original',
                    'Tech Health Survey Adjusted',
                    'Tech Health Variance',
                    'Critical to Business',
                    'User Satisfaction',
                    'Performance Rating',
                    'Strategic Importance',
                    'Survey Response Count',
                    'Overall Consensus Score',
                    'High Variance Flag',
                    'Qualitative Feedback'
                ]
                export_cols = [col for col in main_cols if col in merged_df.columns]
                merged_df[export_cols].to_excel(writer, sheet_name='Survey_Analysis', index=False)

                # Sheet 2: High variance applications
                if 'High Variance Flag' in merged_df.columns:
                    high_var = merged_df[merged_df['High Variance Flag'] == 1]
                    if len(high_var) > 0:
                        var_cols = ['Application Name'] + [col for col in merged_df.columns if 'Variance' in col and 'Flag' not in col]
                        high_var[var_cols].to_excel(writer, sheet_name='High_Variance', index=False)

                # Sheet 3: Survey impact summary
                impact = self.calculate_survey_impact(merged_df)
                impact_data = []

                if 'variance_summary' in impact:
                    impact_data.append(['VARIANCE SUMMARY', ''])
                    for metric, values in impact['variance_summary'].items():
                        impact_data.append([metric, ''])
                        for stat, value in values.items():
                            impact_data.append([f'  {stat}', f'{value:.2f}'])

                if 'consensus_summary' in impact:
                    impact_data.append(['', ''])
                    impact_data.append(['CONSENSUS SUMMARY', ''])
                    for metric, value in impact['consensus_summary'].items():
                        impact_data.append([metric, str(value)])

                if 'sentiment_analysis' in impact:
                    impact_data.append(['', ''])
                    impact_data.append(['SENTIMENT ANALYSIS', ''])
                    for category, count in impact['sentiment_analysis'].items():
                        impact_data.append([category, str(count)])

                impact_df = pd.DataFrame(impact_data, columns=['Metric', 'Value'])
                impact_df.to_excel(writer, sheet_name='Impact_Summary', index=False, header=False)

                # Sheet 4: Needs attention (high criticality, low satisfaction)
                if 'needs_attention' in impact and len(impact['needs_attention']) > 0:
                    needs_df = pd.DataFrame(impact['needs_attention'])
                    needs_df.to_excel(writer, sheet_name='Needs_Attention', index=False)

                # Sheet 5: Qualitative feedback by application
                if 'Qualitative Feedback' in merged_df.columns:
                    feedback_df = merged_df[merged_df['Qualitative Feedback'].notna()][
                        ['Application Name', 'Survey Response Count', 'Qualitative Feedback']
                    ]
                    if len(feedback_df) > 0:
                        feedback_df.to_excel(writer, sheet_name='Qualitative_Feedback', index=False)

            # Apply formatting
            workbook = load_workbook(output_path)

            # Format all sheets
            for sheet_name in workbook.sheetnames:
                ws = workbook[sheet_name]

                # Header formatting
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")

                    # Auto-adjust column widths
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 100)
                        ws.column_dimensions[column_letter].width = adjusted_width

            workbook.save(output_path)

            logger.info(f"Survey analysis export saved to: {output_path}")
            return output_path

        except Exception as e:
            logger.error(f"Error creating survey analysis export {output_path}: {e}")
            raise
